#!/usr/bin/env python3
"""
Excel to JSON Converter for GitHub Actions

Converts an Excel table to JSON format for static site deployment.
Reads from /data/QA.xlsx (worksheet "Tools", table "Tools") and
outputs to /data/tools.json as a flat array of objects.

Author: Automated via speckit.implement
Date: 2025-10-30
"""

import json
import sys
import logging
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.worksheet.table import Table
except ImportError:
    print("ERROR: openpyxl is not installed. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


def parse_arguments() -> tuple[Path, Path]:
    """
    Parse command-line arguments for file paths.
    
    Returns:
        tuple: (excel_path, json_path)
    """
    # Default paths relative to repository root
    excel_path = Path("data/QA.xlsx")
    json_path = Path("data/tools.json")
    
    # Future: Could add argparse for custom paths
    # For now, use defaults as specified in requirements
    
    return excel_path, json_path


def load_excel_file(excel_path: Path) -> Any:
    """
    Load Excel file with error handling.
    
    Args:
        excel_path: Path to the Excel file
        
    Returns:
        Workbook object
        
    Raises:
        FileNotFoundError: If Excel file doesn't exist
        Exception: If file is corrupted or unreadable
    """
    if not excel_path.exists():
        logger.error(f"❌ Excel file not found: {excel_path}")
        logger.error(f"   Current working directory: {Path.cwd()}")
        raise FileNotFoundError(f"Excel file does not exist: {excel_path}")
    
    try:
        logger.info(f"Loading Excel file: {excel_path}")
        # data_only=True extracts calculated values from formulas
        workbook = load_workbook(excel_path, data_only=True)
        logger.info("✓ Excel file loaded successfully")
        return workbook
    except Exception as e:
        logger.error(f"❌ Excel file is corrupted or unreadable")
        logger.error(f"   Error details: {type(e).__name__}: {e}")
        raise Exception(f"Excel file is corrupted or unreadable: {e}")


def validate_worksheet(workbook: Any, worksheet_name: str) -> Worksheet:
    """
    Validate that the specified worksheet exists.
    
    Args:
        workbook: Openpyxl workbook object
        worksheet_name: Name of the worksheet to find
        
    Returns:
        Worksheet object
        
    Raises:
        KeyError: If worksheet doesn't exist
    """
    if worksheet_name not in workbook.sheetnames:
        available_sheets = ", ".join(workbook.sheetnames)
        logger.error(f"❌ Worksheet '{worksheet_name}' not found")
        logger.error(f"   Available worksheets: {available_sheets}")
        raise KeyError(f"Worksheet '{worksheet_name}' not found in Excel file. Available: {available_sheets}")
    
    worksheet = workbook[worksheet_name]
    logger.info(f"✓ Worksheet '{worksheet_name}' found and validated")
    return worksheet


def extract_table(worksheet: Worksheet, table_name: str) -> Table:
    """
    Extract the specified table from the worksheet.
    
    Args:
        worksheet: Openpyxl worksheet object
        table_name: Name of the table to extract
        
    Returns:
        Table object
        
    Raises:
        KeyError: If table doesn't exist
    """
    if table_name not in worksheet.tables:
        available_tables = ", ".join(worksheet.tables.keys()) if worksheet.tables else "None"
        logger.error(f"❌ Table '{table_name}' not found in worksheet")
        logger.error(f"   Available tables: {available_tables}")
        raise KeyError(f"Table '{table_name}' not found in worksheet. Available: {available_tables}")
    
    table = worksheet.tables[table_name]
    logger.info(f"✓ Table '{table_name}' found: {table.ref}")
    return table


def parse_table_headers(worksheet: Worksheet, table: Table) -> List[str]:
    """
    Parse table headers from the first row.
    
    Args:
        worksheet: Openpyxl worksheet object
        table: Table object
        
    Returns:
        List of header names (preserved with spaces and special characters)
    """
    # Get table range
    table_range = worksheet[table.ref]
    
    # First row contains headers
    header_row = table_range[0]
    headers = [cell.value for cell in header_row]
    
    logger.info(f"Parsed {len(headers)} headers: {headers}")
    return headers


def convert_cell_value(value: Any) -> Any:
    """
    Convert Excel cell value to appropriate JSON type.
    
    Args:
        value: Raw cell value from openpyxl
        
    Returns:
        JSON-compatible value (string, number, boolean, null)
    """
    if value is None:
        return None
    
    # Convert datetime to ISO 8601 string
    if isinstance(value, datetime):
        return value.isoformat()
    
    # Preserve strings, numbers, and booleans as-is
    if isinstance(value, (str, int, float, bool)):
        return value
    
    # Fallback: convert to string
    return str(value)


def extract_table_data(worksheet: Worksheet, table: Table) -> List[Dict[str, Any]]:
    """
    Extract all data rows from the table and convert to list of dictionaries.
    
    Args:
        worksheet: Openpyxl worksheet object
        table: Table object
        
    Returns:
        List of dictionaries with headers as keys and cell values
    """
    # Get table range
    table_range = worksheet[table.ref]
    
    # Parse headers from first row
    headers = parse_table_headers(worksheet, table)
    
    # Process data rows (skip header row)
    data_rows = []
    row_count = 0
    
    for row in table_range[1:]:  # Skip header row
        # Create dictionary for this row
        row_dict = {}
        for header, cell in zip(headers, row):
            # Convert cell value to appropriate JSON type
            row_dict[header] = convert_cell_value(cell.value)
        
        data_rows.append(row_dict)
        row_count += 1
    
    logger.info(f"Extracted {row_count} data rows from table")
    return data_rows


def write_json_file(data: List[Dict[str, Any]], json_path: Path) -> None:
    """
    Write data to JSON file with proper formatting.
    
    Args:
        data: List of dictionaries to write
        json_path: Output file path
        
    Raises:
        IOError: If file cannot be written
    """
    try:
        # Ensure parent directory exists
        json_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Write JSON with UTF-8 encoding, 2-space indent, preserve Unicode
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"JSON file written successfully: {json_path}")
        logger.info(f"File size: {json_path.stat().st_size} bytes")
    except Exception as e:
        logger.error(f"Failed to write JSON file: {e}")
        raise IOError(f"Cannot write JSON file: {e}")


def main():
    """Main execution function."""
    logger.info("Starting Excel to JSON conversion")
    
    try:
        # Parse arguments
        excel_path, json_path = parse_arguments()
        logger.info(f"Excel source: {excel_path}")
        logger.info(f"JSON output: {json_path}")
        
        # Load Excel file
        workbook = load_excel_file(excel_path)
        
        # Validate worksheet exists
        worksheet = validate_worksheet(workbook, "Tools")
        
        # Extract table
        table = extract_table(worksheet, "Tools")
        
        # Extract data from table
        data = extract_table_data(worksheet, table)
        
        # Write JSON output
        write_json_file(data, json_path)
        
        logger.info("✓ Conversion completed successfully")
        return 0
        
    except FileNotFoundError as e:
        logger.error(f"❌ File not found error: {e}")
        logger.error("   Please ensure data/QA.xlsx exists in the repository")
        return 1
    except KeyError as e:
        logger.error(f"❌ Missing required element: {e}")
        logger.error("   Please check that the Excel file has:")
        logger.error("   - A worksheet named 'Tools'")
        logger.error("   - A table named 'Tools' in that worksheet")
        return 1
    except Exception as e:
        logger.error(f"❌ Unexpected error: {type(e).__name__}: {e}")
        logger.error("   This may indicate a corrupted Excel file or system issue")
        return 1


if __name__ == "__main__":
    sys.exit(main())
